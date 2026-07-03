from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd
import requests

from ..core.io import atomic_write_csv, safe_extract_zip
from ..core.manifest import write_manifest
from ..core.paths import get_paths


SOURCE_ID = "abs"
SOURCE_URL = "https://www.abs.gov.au/census/find-census-data/datapacks"
DATAPACK_DOWNLOAD_BASE = "https://www.abs.gov.au/census/find-census-data/datapacks/download"

DEFAULT_CENSUS_YEAR = 2021
DEFAULT_STATE = "NSW"
GEOGRAPHY = "POA"

# 2021Census_G01_NSW_POA.csv, 2021Census_G04A_NSW_POA.csv, ...
TABLE_CSV_RE = re.compile(r"^\d{4}Census_(?P<table>G\w+)_(?P<state>[A-Z]+)_(?P<geography>[A-Z]+)\.csv$")
POA_CODE_RE = re.compile(r"^POA(\d+)$")


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/126.0 Safari/537.36"
            ),
            "Referer": SOURCE_URL,
            "Accept": "application/zip,application/octet-stream,*/*",
        }
    )
    return session


def datapack_zip_url(census_year: int | str, state: str, geography: str = GEOGRAPHY) -> str:
    return f"{DATAPACK_DOWNLOAD_BASE}/{census_year}_GCP_{geography}_for_{state}_short-header.zip"


def raw_datapack_zip_path(
    data_dir: str | Path | None, census_year: int | str, state: str, geography: str = GEOGRAPHY
) -> Path:
    return (
        get_paths(data_dir).raw_source_dir(SOURCE_ID)
        / "datapacks"
        / f"{census_year}_GCP_{geography}_for_{state}_short-header.zip"
    )


def raw_extract_dir(
    data_dir: str | Path | None, census_year: int | str, state: str, geography: str = GEOGRAPHY
) -> Path:
    return get_paths(data_dir).raw_source_dir(SOURCE_ID) / str(census_year) / geography / state


def abs_interim_dir(
    data_dir: str | Path | None, census_year: int | str, state: str, geography: str = GEOGRAPHY
) -> Path:
    return get_paths(data_dir).interim_source_dir(SOURCE_ID) / str(census_year) / f"{state}_{geography}"


def abs_poa_partition_path(data_dir: str | Path | None, census_year: int | str) -> Path:
    return get_paths(data_dir).abs_poa_dir / f"census_year={census_year}.csv"


def abs_poa_columns_path(data_dir: str | Path | None, census_year: int | str) -> Path:
    return get_paths(data_dir).abs_poa_dir / f"census_year={census_year}_columns.csv"


def poa_code_to_postcode(code: object) -> str:
    match = POA_CODE_RE.fullmatch(str(code).strip())
    if not match:
        raise ValueError(f"Unexpected POA code format: {code!r}")
    return match.group(1).zfill(4)


def discover_table_csvs(extract_dir: Path, state: str, geography: str = GEOGRAPHY) -> list[Path]:
    suffix = f"_{state}_{geography}.csv"
    return sorted(path for path in extract_dir.rglob(f"*{suffix}") if TABLE_CSV_RE.match(path.name))


def find_metadata_workbook(extract_dir: Path) -> Path:
    candidates = sorted(extract_dir.rglob("Metadata_*.xlsx"))
    if not candidates:
        raise RuntimeError(f"No metadata workbook (Metadata_*.xlsx) found under {extract_dir}")
    return candidates[0]


def _header_and_rows(worksheet, first_cell: str) -> tuple[list, list[tuple]]:
    rows = list(worksheet.iter_rows(values_only=True))
    for index, row in enumerate(rows):
        if row and row[0] == first_cell:
            header = list(row)
            while header and header[-1] is None:
                header.pop()
            data = [row[: len(header)] for row in rows[index + 1 :]]
            return header, data
    raise RuntimeError(f"Could not find a header row starting with {first_cell!r}")


def load_cell_descriptors(metadata_path: Path) -> pd.DataFrame:
    """Short-code -> long-label decode table, keyed by (DataPackfile, Short)."""

    workbook = openpyxl.load_workbook(metadata_path, read_only=True, data_only=True)
    header, data = _header_and_rows(workbook["Cell Descriptors Information"], "Sequential")
    frame = pd.DataFrame(data, columns=header)
    return frame.dropna(subset=["Short", "DataPackfile"])[["Short", "Long", "DataPackfile", "Profiletable"]]


def load_table_names(metadata_path: Path) -> pd.DataFrame:
    """Table code -> table name/population, e.g. G01 -> 'Selected Person Characteristics by Sex'."""

    workbook = openpyxl.load_workbook(metadata_path, read_only=True, data_only=True)
    header, data = _header_and_rows(workbook["Table Number, Name, Population"], "Table Number")
    frame = pd.DataFrame(data, columns=[str(name).strip() for name in header])
    return frame.dropna(subset=["Table Number"])


def pull_abs(
    data_dir: str | Path | None = None,
    census_year: int | str = DEFAULT_CENSUS_YEAR,
    state: str = DEFAULT_STATE,
    geography: str = GEOGRAPHY,
    dry_run: bool = False,
    force: bool = False,
    session: requests.Session | None = None,
) -> pd.DataFrame:
    paths = get_paths(data_dir)
    paths.ensure_base_dirs()
    zip_path = raw_datapack_zip_path(data_dir, census_year, state, geography)
    extract_dir = raw_extract_dir(data_dir, census_year, state, geography)
    url = datapack_zip_url(census_year, state, geography)

    already_extracted = extract_dir.exists() and any(extract_dir.rglob("*.csv"))
    if already_extracted and not force:
        return pd.DataFrame(
            [{"census_year": census_year, "state": state, "geography": geography, "url": url, "path": str(extract_dir), "status": "exists"}]
        )
    if dry_run:
        return pd.DataFrame(
            [{"census_year": census_year, "state": state, "geography": geography, "url": url, "path": str(extract_dir), "status": "pending"}]
        )

    zip_path.parent.mkdir(parents=True, exist_ok=True)
    session = session or make_session()
    print(f"Downloading ABS {geography} GCP DataPack for {state} {census_year}: {url}")
    response = session.get(url, timeout=180)
    response.raise_for_status()
    zip_path.write_bytes(response.content)
    safe_extract_zip(zip_path, extract_dir)
    return pd.DataFrame(
        [{"census_year": census_year, "state": state, "geography": geography, "url": url, "path": str(extract_dir), "status": "downloaded"}]
    )


def extract_abs(
    data_dir: str | Path | None = None,
    census_year: int | str = DEFAULT_CENSUS_YEAR,
    state: str = DEFAULT_STATE,
    geography: str = GEOGRAPHY,
) -> pd.DataFrame:
    paths = get_paths(data_dir)
    paths.ensure_base_dirs()
    extract_dir = raw_extract_dir(data_dir, census_year, state, geography)
    if not extract_dir.exists():
        raise RuntimeError(f"Missing raw ABS DataPack extract at {extract_dir}; run `abs pull` first.")

    table_csvs = discover_table_csvs(extract_dir, state, geography)
    if not table_csvs:
        raise RuntimeError(f"No {geography} table CSVs found under {extract_dir}; DataPack layout may have changed.")

    geo_code_column = f"{geography}_CODE_{census_year}"
    interim_dir = abs_interim_dir(data_dir, census_year, state, geography)
    interim_dir.mkdir(parents=True, exist_ok=True)

    statuses = []
    for csv_path in table_csvs:
        match = TABLE_CSV_RE.match(csv_path.name)
        table = match.group("table")
        frame = pd.read_csv(csv_path, dtype=str)
        if frame.empty:
            raise RuntimeError(f"ZERO rows in {csv_path}; investigate DataPack layout.")
        if geo_code_column not in frame.columns:
            raise RuntimeError(f"Expected column {geo_code_column!r} not found in {csv_path}")
        frame["postcode"] = frame[geo_code_column].map(poa_code_to_postcode)
        frame = frame.drop(columns=[geo_code_column])
        output_path = atomic_write_csv(frame, interim_dir / f"{table}.csv")
        statuses.append(
            {"table": table, "rows": int(frame.shape[0]), "columns": int(frame.shape[1] - 1), "path": str(output_path), "status": "written"}
        )
    return pd.DataFrame(statuses)


def transform_abs(
    data_dir: str | Path | None = None,
    census_year: int | str = DEFAULT_CENSUS_YEAR,
    state: str = DEFAULT_STATE,
    geography: str = GEOGRAPHY,
) -> pd.DataFrame:
    paths = get_paths(data_dir)
    paths.ensure_base_dirs()
    interim_dir = abs_interim_dir(data_dir, census_year, state, geography)
    interim_csvs = sorted(interim_dir.glob("G*.csv"))
    if not interim_csvs:
        raise RuntimeError(f"No extracted ABS tables found under {interim_dir}; run `abs extract` first.")

    metadata_path = find_metadata_workbook(raw_extract_dir(data_dir, census_year, state, geography))
    cell_descriptors = load_cell_descriptors(metadata_path)
    table_names = load_table_names(metadata_path)

    frames = []
    column_rows = []
    for csv_path in interim_csvs:
        table = csv_path.stem
        frame = pd.read_csv(csv_path, dtype=str).set_index("postcode")
        rename = {short_code: f"{table}__{short_code}" for short_code in frame.columns}
        for short_code, wide_column in rename.items():
            column_rows.append({"column": wide_column, "table": table, "short_code": short_code})
        frames.append(frame.rename(columns=rename))

    # Every table shares the same postcode set, but concat(join="outer") tolerates
    # a table that's short a few suppressed/unavailable geographies.
    wide = pd.concat(frames, axis=1, join="outer").reset_index()
    if wide.empty:
        raise RuntimeError(f"ZERO rows produced for ABS {geography} {census_year}; investigate.")
    wide.insert(1, "census_year", census_year)

    output_path = atomic_write_csv(wide, abs_poa_partition_path(data_dir, census_year))

    columns_frame = pd.DataFrame(column_rows, columns=["column", "table", "short_code"])
    columns_frame = columns_frame.merge(
        cell_descriptors, left_on=["table", "short_code"], right_on=["DataPackfile", "Short"], how="left"
    )
    columns_frame = columns_frame.merge(table_names, left_on="table", right_on="Table Number", how="left")
    columns_frame = columns_frame[["column", "table", "short_code", "Long", "Table Name"]].rename(
        columns={"Long": "long_label", "Table Name": "table_name"}
    )
    atomic_write_csv(columns_frame, abs_poa_columns_path(data_dir, census_year))

    refresh_abs_poa_manifest(data_dir=data_dir)
    return pd.DataFrame(
        [{"census_year": census_year, "path": str(output_path), "rows": int(wide.shape[0]), "columns": int(wide.shape[1]), "status": "written"}]
    )


def discover_abs_poa_partitions(data_dir: str | Path | None = None) -> list[tuple[Path, str, str]]:
    paths = get_paths(data_dir)
    partitions = []
    for path in paths.abs_poa_dir.glob("census_year=*.csv"):
        match = re.fullmatch(r"census_year=(\d{4})\.csv", path.name)
        if not match:
            continue
        year = match.group(1)
        partitions.append((path, f"{year}-01-01", f"{year}-12-31"))
    return partitions


def refresh_abs_poa_manifest(data_dir: str | Path | None = None) -> pd.DataFrame:
    paths = get_paths(data_dir)
    return write_manifest(
        data_dir=paths.data_dir,
        manifest_path=paths.abs_poa_manifest,
        source=SOURCE_ID,
        dataset="poa_gcp",
        partitions=discover_abs_poa_partitions(data_dir=data_dir),
    )


def update_abs(
    data_dir: str | Path | None = None,
    census_year: int | str = DEFAULT_CENSUS_YEAR,
    state: str = DEFAULT_STATE,
    geography: str = GEOGRAPHY,
    dry_run: bool = False,
    force: bool = False,
) -> pd.DataFrame:
    pull_abs(data_dir=data_dir, census_year=census_year, state=state, geography=geography, dry_run=dry_run, force=force)
    if dry_run:
        return pd.DataFrame()
    extract_abs(data_dir=data_dir, census_year=census_year, state=state, geography=geography)
    return transform_abs(data_dir=data_dir, census_year=census_year, state=state, geography=geography)
